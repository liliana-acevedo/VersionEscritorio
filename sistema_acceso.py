import customtkinter as ctk
from cliente_supabase import supabase
import tkinter as tk
from datetime import datetime, timedelta
import threading
import pandas as pd
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.styles.borders import BORDER_THIN
import os
from PIL import Image as PILImage 

# IMPORTACIÓN DEL CONTROLADOR DE GRÁFICOS
try:
    import matplotlib
    matplotlib.use('TkAgg') 
    from controladores_graficos import mostrar_pantalla_graficos
except ImportError as e:
    print(f"Error/Advertencia Matplotlib o graficos: {e}")
    def mostrar_pantalla_graficos(root, callback_volver):
        messagebox.showerror("Error", "No se pudo cargar el módulo de gráficos.")

# Referencias Globales de la Interfaz 
app_root = None

# Referencias para la Pantalla 'Agregar Usuario' 
registro_entries = {}
registro_notificacion = None

# Referencias para la Pantalla 'Agregar Departamento'
depto_entry = None
depto_notificacion = None

# Variable global para usuario seleccionado
usuario_seleccionado = None

# Funciones de Utilidad
def _clear_widgets(root):
    for widget in root.winfo_children():
        widget.destroy()

def formatear_fecha(iso_fecha_str):
    if not iso_fecha_str:
        return "Pendiente"
    try:
        dt_obj = datetime.fromisoformat(iso_fecha_str)
        return dt_obj.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(iso_fecha_str)

# --- LÓGICA DE SESIÓN ---
def cerrar_sesion(root):
    # Importamos setup_login_app aquí para evitar el error "Circular Import"
    # ya que login.py importa sistema_acceso.py
    from login import setup_login_app
    setup_login_app(root)

#  Carga de Datos Estructurales (Departamentos y Roles)
def obtener_departamentos():
    departamentos_map = {}
    try:
        resp = supabase.table("Departamento").select("id_departamento, nombre_departamento").execute()
        for item in resp.data or []:
            nombre = item.get("nombre_departamento") or ""
            idd = item.get("id_departamento")
            if nombre and idd is not None:
                departamentos_map[str(nombre)] = idd
    except Exception as e:
        print("Error al obtener departamentos:", e)
    return departamentos_map

def obtener_roles():
    roles_map = {}
    try:
        resp = supabase.table("Rol").select("id_rol, nombre_rol").execute()
        for item in resp.data or []:
            nombre = item.get("nombre_rol") or ""
            idd = item.get("id_rol")
            if nombre and idd is not None:
                roles_map[str(nombre)] = idd
    except Exception as e:
        print("Error al obtener roles:", e)
    return roles_map

def obtener_usuarios_completos():
    try:
        response = (
            supabase.table('Usuario')
            .select('nombre, apellido, cedula, Departamento(nombre_departamento), Rol(nombre_rol)')
            .execute()
        )
        datos = response.data

        if not datos:
            return pd.DataFrame(columns=['nombre', 'apellido', 'cedula', 'departamento', 'rol'])

        usuarios_procesados = []
        for usuario in datos:
            usuario_procesado = {
                'nombre': usuario.get('nombre', ''),
                'apellido': usuario.get('apellido', ''),
                'cedula': usuario.get('cedula', ''),
                'departamento': 'Sin departamento',
                'rol': 'Sin rol'
            }
            
            depto_data = usuario.get('Departamento')
            if depto_data and isinstance(depto_data, list) and len(depto_data) > 0:
                usuario_procesado['departamento'] = depto_data[0].get('nombre_departamento', 'Sin departamento')
            elif depto_data and isinstance(depto_data, dict):
                usuario_procesado['departamento'] = depto_data.get('nombre_departamento', 'Sin departamento')
            
            rol_data = usuario.get('Rol')
            if rol_data and isinstance(rol_data, list) and len(rol_data) > 0:
                usuario_procesado['rol'] = rol_data[0].get('nombre_rol', 'Sin rol')
            elif rol_data and isinstance(rol_data, dict):
                usuario_procesado['rol'] = rol_data.get('nombre_rol', 'Sin rol')
                
            usuarios_procesados.append(usuario_procesado)

        df_usuarios = pd.DataFrame(usuarios_procesados)
        return df_usuarios

    except Exception as e:
        print(f"Ocurrió un error al obtener datos de Supabase: {e}")
        return pd.DataFrame(columns=['nombre', 'apellido', 'cedula', 'departamento', 'rol'])

# FUNCIONES PARA ELIMINAR Y EDITAR USUARIOS
def eliminar_usuario(cedula, nombre_completo, row_frame=None):
    def _eliminar():
        try:
            cedula_int = int(cedula)
            response = supabase.table("Usuario").delete().eq("cedula", cedula_int).execute()
            if response.data:
                print(f"Usuario {nombre_completo} eliminado correctamente")
                def eliminar_fila_ui():
                    if row_frame and row_frame.winfo_exists():
                        row_frame.destroy()
                    global usuario_seleccionado
                    if usuario_seleccionado and usuario_seleccionado['cedula'] == cedula:
                        usuario_seleccionado = None
                        # Limpiar etiquetas de selección si existen
                        for widget in app_root.winfo_children():
                            if isinstance(widget, ctk.CTkFrame):
                                for child in widget.winfo_children():
                                    if isinstance(child, ctk.CTkFrame):
                                        for subchild in child.winfo_children():
                                            if hasattr(subchild, 'cget') and "USUARIO SELECCIONADO" in subchild.cget("text", "").upper():
                                                subchild.configure(text="NINGÚN USUARIO SELECCIONADO", text_color="white")
                app_root.after(0, eliminar_fila_ui)
            else:
                app_root.after(0, lambda: messagebox.showerror("Error", f"No se pudo eliminar al usuario {nombre_completo}"))
        except Exception as e:
            app_root.after(0, lambda: messagebox.showerror("Error", f"Error al eliminar usuario: {e}"))
    
    confirmar = tk.messagebox.askyesno(
        "Confirmar Eliminación", 
        f"¿Está seguro de que desea eliminar al usuario:\n{nombre_completo}?\n\nCédula: {cedula}"
    )
    if confirmar:
        threading.Thread(target=_eliminar, daemon=True).start()

# Manejo Seguro de Notificaciones
def _set_registro_notificacion(text, color):
    global registro_notificacion, app_root
    if not registro_notificacion or not app_root:
        return
    app_root.after(0, lambda: registro_notificacion.configure(text=text, text_color=color))

def _set_depto_notificacion(text, color):
    global depto_notificacion, app_root
    if not depto_notificacion or not app_root:
        return
    app_root.after(0, lambda: depto_notificacion.configure(text=text, text_color=color))

def _clear_registro_campos():
    global registro_entries, app_root
    if not app_root:
        return
    def _clear():
        for k in ['cedula', 'nombre', 'apellido']:
            ent = registro_entries.get(k)
            if ent:
                try:
                    ent.delete(0, 'end')
                except Exception:
                    pass
    app_root.after(0, _clear)

def abrir_ventana_seleccion_depto(root, display_entry, nombre_var):
    deptos_map = obtener_departamentos()
    all_deptos = sorted(list(deptos_map.keys())) 

    ventana = ctk.CTkToplevel(root)
    ventana.title("Seleccionar Departamento")
    ventana.configure(fg_color="#F7F9FB")
    ventana.grab_set()
    ventana.focus_force()
    ventana.geometry("500x500") 
    ventana.resizable(False, False) 
    
    contenido = ctk.CTkFrame(ventana, fg_color="#FFFFFF")
    contenido.pack(padx=20, pady=20, fill="both", expand=True) 
    contenido.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(contenido, text="Buscar Departamento", font=ctk.CTkFont(size=18, weight="bold"), text_color="#0C4A6E").grid(row=0, column=0, pady=(10, 15), sticky="w")
    
    search_entry = ctk.CTkEntry(contenido, placeholder_text="Escriba para buscar...", width=450, height=35)
    search_entry.grid(row=1, column=0, pady=(0, 10), sticky="ew")
    
    scroll_frame = ctk.CTkScrollableFrame(contenido, fg_color="#F9FAFB")
    scroll_frame.grid(row=2, column=0, sticky="nsew", pady=(0, 0)) 
    scroll_frame.grid_columnconfigure(0, weight=1)

    contenido.grid_rowconfigure(2, weight=1)

    def seleccionar_depto(nombre):
        display_entry.configure(state="normal")
        display_entry.delete(0, 'end')
        display_entry.insert(0, nombre)
        display_entry.configure(state="readonly")
        nombre_var.set(nombre)
        ventana.destroy()

    def render_list(filtro=""):
        for widget in scroll_frame.winfo_children():
            widget.destroy()
        
        filtro_lower = filtro.lower().strip()
        deptos_filtrados = []
        
        if filtro_lower:
            deptos_filtrados = [nombre for nombre in all_deptos if nombre.lower().startswith(filtro_lower)]
        else:
            deptos_filtrados = all_deptos
    
        if not deptos_filtrados:
            lbl = ctk.CTkLabel(scroll_frame, text="No se encontraron departamentos", text_color="#6B7280", font=ctk.CTkFont(size=12))
            lbl.grid(row=0, column=0, sticky="ew", pady=10)
            return
        
        for i, nombre in enumerate(deptos_filtrados):
            btn = ctk.CTkButton(
                scroll_frame, text=nombre, fg_color="transparent", hover_color="#E0F2FE", 
                text_color="black", corner_radius=0, anchor="w",
                command=lambda n=nombre: seleccionar_depto(n)
            )
            btn.grid(row=i, column=0, sticky="ew", pady=(1, 1))

    def filtrar_lista(event=None):
        texto_busqueda = search_entry.get()
        render_list(texto_busqueda)
        
    search_entry.bind("<KeyRelease>", filtrar_lista)
    search_entry.focus_set()
    render_list()
    
    ctk.CTkButton(contenido, text="CANCELAR", fg_color="#6B7280", hover_color="#4B5563", width=150, height=35, command=ventana.destroy).grid(row=3, column=0, pady=(10, 0))


# PANTALLA: AGREGAR NUEVO DEPARTAMENTO
def mostrar_pantalla_departamentos(root):
    global depto_entry, depto_notificacion

    _clear_widgets(root)
    root.title("GESTIÓN DE DEPARTAMENTOS")

    main = ctk.CTkFrame(root, fg_color="#F2F5F9")
    main.pack(expand=True, fill="both")
    main.grid_rowconfigure(1, weight=1)
    main.grid_columnconfigure(0, weight=1)

    header = ctk.CTkFrame(main, fg_color="#0C4A6E", height=70, corner_radius=0)
    header.grid(row=0, column=0, sticky="ew")
    header.grid_columnconfigure(0, weight=1) 
    header.grid_columnconfigure(1, weight=0)  

    ctk.CTkLabel(header, text="Gestión de Departamentos", text_color="white", font=ctk.CTkFont(size=20, weight="bold")).grid(row=0, column=0, padx=20, pady=18, sticky="w")

    try:
        ruta_volver = os.path.join("imagen", "volver.png")
        icono_volver = ctk.CTkImage(light_image=PILImage.open(ruta_volver), size=(20, 20))
        texto_btn = "" 
        ancho_btn = 50 
    except Exception:
        icono_volver = None
        texto_btn = "VOLVER" 
        ancho_btn = 90

    volver_btn = ctk.CTkButton(header, text=texto_btn, image=icono_volver, width=ancho_btn, height=36, fg_color="#3D89D1", hover_color="#1E3D8F", command=lambda: mostrar_pantalla_principal(root))
    volver_btn.grid(row=0, column=1, padx=20, pady=17, sticky="e")

    content = ctk.CTkFrame(main, fg_color="transparent")
    content.grid(row=1, column=0, sticky="nsew", padx=12, pady=12)
    content.grid_columnconfigure(0, weight=3)
    content.grid_columnconfigure(1, weight=2)
    content.grid_rowconfigure(0, weight=1)

    left = ctk.CTkFrame(content, fg_color="#FFFFFF", corner_radius=10)
    left.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
    left.grid_rowconfigure(1, weight=1)
    left.grid_columnconfigure(0, weight=1)

    actions = ctk.CTkFrame(left, fg_color="transparent", height=50) 
    actions.grid_propagate(False) 
    actions.grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 0))
    actions.grid_columnconfigure(1, weight=1)

    try:
        ruta_eliminar = os.path.join("imagen", "eliminar.png")
        icono_eliminar = ctk.CTkImage(light_image=PILImage.open(ruta_eliminar), size=(20, 20))
        texto_eliminar = ""
        ancho_eliminar = 40 
    except Exception:
        icono_eliminar = None
        texto_eliminar = "ELIMINAR"
        ancho_eliminar = 110

    eliminar_btn = ctk.CTkButton(actions, text=texto_eliminar, image=icono_eliminar, fg_color="#DC2626", hover_color="#B91C1C", width=ancho_eliminar, height=34, command=lambda: on_eliminar())
    eliminar_btn.grid(row=0, column=0)

    selection_label = ctk.CTkLabel(actions, text="NINGÚN DEPARTAMENTO SELECCIONADO", text_color="white", fg_color="#0C4A6E", corner_radius=6, font=ctk.CTkFont(size=11, weight="bold"), padx=10, pady=5)
    selection_label.grid(row=0, column=1, sticky="e", padx=(10, 0))

    rows = ctk.CTkScrollableFrame(left, fg_color="transparent")
    rows.grid(row=1, column=0, sticky="nsew", padx=12, pady=12)
    rows.grid_columnconfigure(0, weight=1)

    selected = {"id": None, "nombre": None}

    right = ctk.CTkFrame(content, fg_color="#FFFFFF", corner_radius=10)
    right.grid(row=0, column=1, sticky="nsew")
    right.grid_columnconfigure(0, weight=1)

    ctk.CTkLabel(right, text="AGREGAR / EDITAR DEPARTAMENTO", font=ctk.CTkFont(size=18, weight="bold"), text_color="#1E3D8F").grid(row=0, column=0, pady=(30, 15), padx=20)

    depto_entry = ctk.CTkEntry(right, placeholder_text="Nombre del departamento", width=450, height=45, font=ctk.CTkFont(size=14), border_color="#CDCECF")
    depto_entry.grid(row=1, column=0, pady=(5, 5), padx=20)

    depto_notificacion = ctk.CTkLabel(right, text="", text_color="#16A34A")
    depto_notificacion.grid(row=2, column=0, pady=(0, 20))

    def on_guardar():
        nombre = depto_entry.get().strip().upper()
        if not nombre:
            depto_notificacion.configure(text="Ingrese un nombre válido.", text_color="orange")
            return

        if selected["id"]:
            actualizar_departamento(selected["id"], nombre)
            cargar_departamentos()
            limpiar_seleccion()
        else:
            try:
                supabase.table("Departamento").insert({"nombre_departamento": nombre}).execute()
                depto_notificacion.configure(text="Departamento agregado con éxito.", text_color="#16A34A")
                cargar_departamentos()
                limpiar_seleccion()
            except Exception as e:
                print(f"ERROR: {e}")
                depto_notificacion.configure(text=f"Error al guardar: {str(e)}", text_color="red")

    guardar_btn = ctk.CTkButton(right, text="GUARDAR CAMBIOS", fg_color="#16A34A", hover_color="#15803D", font=ctk.CTkFont(size=13, weight="bold"), width=450, height=45, command=on_guardar)
    guardar_btn.grid(row=3, column=0, pady=(10, 10), padx=20)

    cancelar_btn = ctk.CTkButton(right, text="CANCELAR", fg_color="#8b8a8a", hover_color="#777373", font=ctk.CTkFont(size=13, weight="bold"), width=450, height=45, command=lambda: limpiar_seleccion())
    cancelar_btn.grid(row=4, column=0, pady=(0, 20))
    
    def limpiar_seleccion():
        selected["id"] = None
        selected["nombre"] = None
        depto_entry.delete(0, "end")
        selection_label.configure(text="Ningún departamento seleccionado")
        guardar_btn.configure(text="GUARDAR")
        depto_notificacion.configure(text="", text_color="#16A34A")

    def seleccionar(id_dep, nombre, frame):
        for r in rows.winfo_children():
            r.configure(fg_color="transparent")

        frame.configure(fg_color="#E0F2FE")
        selected["id"] = id_dep
        selected["nombre"] = nombre
        selection_label.configure(text=f"Seleccionado: {nombre}")
        depto_entry.delete(0, "end")
        depto_entry.insert(0, nombre)
        guardar_btn.configure(text="GUARDAR CAMBIOS")
        depto_notificacion.configure(text="", text_color="#16A34A") 

    def actualizar_departamento(id_dep, nombre):
        try:
            supabase.table("Departamento").update({"nombre_departamento": nombre}).eq("id_departamento", id_dep).execute()
            depto_notificacion.configure(text="Departamento actualizado.", text_color="#16A34A")
        except Exception as e:
            depto_notificacion.configure(text=f"Error: {e}", text_color="red")

    def on_eliminar():
        if not selected["id"]:
            tk.messagebox.showwarning("Advertencia", "Seleccione un departamento.")
            return
        if not tk.messagebox.askyesno("Confirmar", f"¿Eliminar {selected['nombre']}?"):
            return
        try:
            supabase.table("Departamento").delete().eq("id_departamento", selected["id"]).execute()
            depto_notificacion.configure(text="Departamento eliminado.", text_color="#16A34A")
        except Exception as e:
            depto_notificacion.configure(text=f"Error: {e}", text_color="red")
        cargar_departamentos()
        limpiar_seleccion()

    def cargar_departamentos():
        for w in rows.winfo_children():
            w.destroy()
        try:
            data = supabase.table("Departamento").select("id_departamento, nombre_departamento").order("nombre_departamento").execute().data or []
        except:
            data = []
        for d in data:
            f = ctk.CTkFrame(rows, fg_color="transparent", height=42, corner_radius=0)
            f.pack(fill="x", pady=3)
            f.grid_columnconfigure(0, weight=1, uniform="deptos")
            f.configure(width=rows.winfo_width())
            lbl = ctk.CTkLabel(f, text=d["nombre_departamento"], font=ctk.CTkFont(size=14), anchor="w")
            lbl.grid(row=0, column=0, sticky="w", padx=10)
            def on_select(e=None, i=d["id_departamento"], n=d["nombre_departamento"], fr=f):
                for r in rows.winfo_children():
                    r.configure(fg_color="transparent")
                fr.configure(fg_color="#E0F2FE")
                seleccionar(i, n, fr)
            f.bind("<Button-1>", on_select)
            lbl.bind("<Button-1>", on_select)

    cargar_departamentos()


# Mapas de Datos y Lógica de Servicios
def map_usuarios_por_cedula():
    try:
        resp = supabase.table("Usuario").select("cedula, nombre, apellido").execute()
        usuarios = resp.data or []
        mapa = {}
        for u in usuarios:
            ced = u.get("cedula")
            nombre = (u.get("nombre") or "").strip()
            apellido = (u.get("apellido") or "").strip()
            if ced:
                mapa[str(ced)] = f"{nombre} {apellido}".strip() or str(ced)
        return mapa
    except Exception as e:
        print("Error al obtener usuarios:", e)
        return {}

def traducir_estado(valor):
    return {1: "Pendiente", 2: "Completado", 3: "Recibido"}.get(int(valor), "Desconocido") if valor else "Desconocido"

def obtener_servicios_filtrados_base(query_builder):
    try:
        resp = query_builder.execute()
        servicios = resp.data or []
        departamentos = supabase.table("Departamento").select("id_departamento, nombre_departamento").execute().data or []
        dep_map = {str(d["id_departamento"]): d["nombre_departamento"] for d in departamentos}

        for s in servicios:
            dep_val = s.get("departamento")
            if isinstance(dep_val, (int, float)) or (isinstance(dep_val, str) and dep_val.isdigit()):
                s["Departamento"] = dep_map.get(str(dep_val), "Desconocido")
            elif isinstance(dep_val, str):
                s["Departamento"] = dep_val.strip()
            else:
                s["Departamento"] = "Desconocido"
        return servicios
    except Exception as e:
        print("Error obtener servicios:", e)
        return []


# PANTALLA DE REGISTRO DE USUARIO 
def mostrar_pantalla_registro(root):
    global registro_entries, registro_notificacion, app_root, usuario_seleccionado
    app_root = root
    usuario_seleccionado = None
    _clear_widgets(root)
    root.title("Gestión de Usuarios")

    departamentos_map = obtener_departamentos()
    roles_map = obtener_roles()
    departamento_names = list(departamentos_map.keys())
    rol_names = list(roles_map.keys())

    main_frame = ctk.CTkFrame(root, fg_color="#F7F9FB")
    main_frame.pack(expand=True, fill="both")
    main_frame.grid_rowconfigure(1, weight=1)
    main_frame.grid_columnconfigure(0, weight=1)

    header_frame = ctk.CTkFrame(main_frame, fg_color="#0C4A6E", corner_radius=0, height=70)
    header_frame.grid(row=0, column=0, sticky="ew")
    header_frame.grid_columnconfigure(1, weight=1)
    header_frame.grid_columnconfigure(2, weight=0)

    ctk.CTkLabel(header_frame, text="GESTIÓN DE USUARIOS", font=ctk.CTkFont(size=22, weight="bold"), text_color="white").grid(row=0, column=1, padx=(30, 20), pady=15, sticky="w")

    try:
        ruta_volver_reg = os.path.join("imagen", "volver.png")
        icono_volver_reg = ctk.CTkImage(light_image=PILImage.open(ruta_volver_reg), size=(20, 20))
        text_reg = ""
        width_reg = 50
    except Exception:
        icono_volver_reg = None
        text_reg = "VOLVER"
        width_reg = 120

    ctk.CTkButton(header_frame, text=text_reg, image=icono_volver_reg, fg_color="#3D89D1", hover_color="#1E3D8F", corner_radius=8, width=width_reg, height=40, command=lambda: mostrar_pantalla_principal(root)).grid(row=0, column=2, padx=(10, 20), pady=12, sticky="e")

    content_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
    content_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
    content_frame.grid_rowconfigure(0, weight=1)
    content_frame.grid_columnconfigure(0, weight=3)
    content_frame.grid_columnconfigure(1, weight=1)

    col_vacia_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
    col_vacia_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
    
    try:
        df_usuarios = obtener_usuarios_completos() 
        botones_superior_frame = ctk.CTkFrame(col_vacia_frame, fg_color="transparent")
        botones_superior_frame.pack(fill="x", padx=20, pady=(0, 10))
        
        try:
            ruta_eliminar_reg = os.path.join("imagen", "eliminar.png")
            icono_eliminar_reg = ctk.CTkImage(light_image=PILImage.open(ruta_eliminar_reg), size=(20, 20))
            texto_elim_reg = ""
            ancho_elim_reg = 40
        except Exception:
            icono_eliminar_reg = None
            texto_elim_reg = "ELIMINAR"
            ancho_elim_reg = 120

        btn_eliminar_superior = ctk.CTkButton(botones_superior_frame, text=texto_elim_reg, image=icono_eliminar_reg, fg_color="#DC2626", hover_color="#B91C1C", font=ctk.CTkFont(size=13, weight="bold"), width=ancho_elim_reg, height=35, command=lambda: _eliminar_usuario_seleccionado())
        btn_eliminar_superior.pack(side="left", padx=(0, 10))
        
        seleccion_label = ctk.CTkLabel(botones_superior_frame, text="NINGÚN USUARIO SELECCIONADO", text_color="white", fg_color="#0C4A6E", corner_radius=6, font=ctk.CTkFont(size=11, weight="bold"), padx=10, pady=5)
        seleccion_label.pack(side="right", padx=10)

        if df_usuarios.empty:
            ctk.CTkLabel(col_vacia_frame, text="No se encontraron usuarios en la base de datos.", text_color="#1E3D8F", font=ctk.CTkFont(size=14)).pack(pady=10)
        else:
            table_container = ctk.CTkFrame(col_vacia_frame, fg_color="#FFFFFF", corner_radius=10, border_width=1, border_color="#E6E6E6")
            table_container.pack(fill="both", expand=True, padx=20, pady=10)
            
            header_frame_table = ctk.CTkFrame(table_container, fg_color="#F3F4F6", corner_radius=0)
            header_frame_table.pack(fill="x")
            header_frame_table.grid_columnconfigure(0, weight=0, minsize=120)
            header_frame_table.grid_columnconfigure(1, weight=0, minsize=120)
            header_frame_table.grid_columnconfigure(2, weight=0, minsize=100)
            header_frame_table.grid_columnconfigure(3, weight=1, minsize=350)
            header_frame_table.grid_columnconfigure(4, weight=0, minsize=120)
            
            ctk.CTkLabel(header_frame_table, text="NOMBRE", font=ctk.CTkFont(size=13, weight="bold"), text_color="#374151", anchor="w").grid(row=0, column=0, padx=8, pady=10, sticky="w")
            ctk.CTkLabel(header_frame_table, text="APELLIDO", font=ctk.CTkFont(size=13, weight="bold"), text_color="#374151", anchor="w").grid(row=0, column=1, padx=8, pady=10, sticky="w")
            ctk.CTkLabel(header_frame_table, text="CÉDULA", font=ctk.CTkFont(size=13, weight="bold"), text_color="#374151", anchor="w").grid(row=0, column=2, padx=8, pady=10, sticky="w")
            ctk.CTkLabel(header_frame_table, text="DEPARTAMENTO", font=ctk.CTkFont(size=13, weight="bold"), text_color="#374151", anchor="w").grid(row=0, column=3, padx=8, pady=10, sticky="w")
            ctk.CTkLabel(header_frame_table, text="ROL", font=ctk.CTkFont(size=13, weight="bold"), text_color="#374151", anchor="w").grid(row=0, column=4, padx=8, pady=10, sticky="w")

            scroll_frame = ctk.CTkScrollableFrame(table_container, fg_color="#FFFFFF", corner_radius=0)
            scroll_frame.pack(fill="both", expand=True)
            scroll_frame.grid_columnconfigure(0, weight=0, minsize=120)
            scroll_frame.grid_columnconfigure(1, weight=0, minsize=120)
            scroll_frame.grid_columnconfigure(2, weight=0, minsize=100)
            scroll_frame.grid_columnconfigure(3, weight=1, minsize=350)
            scroll_frame.grid_columnconfigure(4, weight=0, minsize=120)
            
            def seleccionar_usuario(cedula, nombre_completo, row_frame, usuario_data):
                global usuario_seleccionado
                for widget in scroll_frame.winfo_children():
                    if isinstance(widget, ctk.CTkFrame):
                        index = scroll_frame.winfo_children().index(widget)
                        bg_color = "#FFFFFF" if index % 2 == 0 else "#F9FAFB"
                        widget.configure(fg_color=bg_color)
                row_frame.configure(fg_color="#E0F2FE")
                usuario_seleccionado = {'cedula': cedula, 'nombre_completo': nombre_completo, 'row_frame': row_frame, 'data': usuario_data}
                seleccion_label.configure(text=f"SELECCIONADO: {nombre_completo.upper()}", text_color="white")
                cargar_datos_formulario(usuario_data)
            
            for i, row in df_usuarios.iterrows():
                bg_color = "#FFFFFF" if i % 2 == 0 else "#F9FAFB" 
                text_color = "#374151"
               
                row_frame = ctk.CTkFrame(scroll_frame, fg_color=bg_color, corner_radius=0, height=35)
                row_frame.pack(fill="x")
                row_frame.grid_columnconfigure(0, weight=0, minsize=120)
                row_frame.grid_columnconfigure(1, weight=0, minsize=120)
                row_frame.grid_columnconfigure(2, weight=0, minsize=100)
                row_frame.grid_columnconfigure(3, weight=1, minsize=350)
                row_frame.grid_columnconfigure(4, weight=0, minsize=120)

                nombre = str(row.get('nombre', '')).strip().upper()
                apellido = str(row.get('apellido', '')).strip().upper()
                cedula = str(row['cedula'])
                departamento = str(row.get('departamento', 'Sin departamento')).strip()
                rol = str(row.get('rol', 'Sin rol')).strip()
                
                if rol.lower() == 'administrador':
                    rol_mostrar = "administrador"
                elif rol.lower() == 'usuario' or rol.lower() == 'usuario estándar':
                    rol_mostrar = "usuario"
                elif rol.lower() == 'tecnico de soporte':
                    rol_mostrar = "tecnico de soporte"
                else:
                    rol_mostrar = rol.lower()
                
                nombre_completo = f"{nombre} {apellido}".strip()
                usuario_data = {'cedula': cedula, 'nombre': nombre, 'apellido': apellido, 'departamento': departamento, 'rol': rol}

                row_frame.bind("<Button-1>", lambda e, c=cedula, n=nombre_completo, rf=row_frame, ud=usuario_data: seleccionar_usuario(c, n, rf, ud))
                
                ctk.CTkLabel(row_frame, text=nombre, font=ctk.CTkFont(size=13), text_color=text_color, anchor="w").grid(row=0, column=0, padx=8, pady=8, sticky="w")
                ctk.CTkLabel(row_frame, text=apellido, font=ctk.CTkFont(size=13), text_color=text_color, anchor="w").grid(row=0, column=1, padx=8, pady=8, sticky="w")
                ctk.CTkLabel(row_frame, text=cedula, font=ctk.CTkFont(size=13), text_color=text_color, anchor="w").grid(row=0, column=2, padx=8, pady=8, sticky="w")
                ctk.CTkLabel(row_frame, text=departamento, font=ctk.CTkFont(size=13), text_color=text_color, anchor="w").grid(row=0, column=3, padx=8, pady=8, sticky="w")
                ctk.CTkLabel(row_frame, text=rol_mostrar, font=ctk.CTkFont(size=13), text_color=text_color, anchor="w").grid(row=0, column=4, padx=8, pady=8, sticky="w")

    except Exception as e:
        ctk.CTkLabel(col_vacia_frame, text=f"Error al cargar usuarios: {e}", text_color="red", font=ctk.CTkFont(size=14)).pack(pady=20, padx=20)

    form_frame = ctk.CTkFrame(content_frame, fg_color="#FFFFFF", corner_radius=10)
    form_frame.grid(row=0, column=1, pady=10, padx=20, ipadx=20, ipady=20, sticky="n")

    ctk.CTkLabel(form_frame, text="FORMULARIO DE USUARIO", font=ctk.CTkFont(size=18, weight="bold"), text_color="#1E3D8F").pack(pady=(10, 20))

    ANCHO_INPUT = 340
    ALTO_INPUT = 40
    COLOR_BORDE = "#94A3B8"
    COLOR_PLACEHOLDER = "#9CA3AF"

    registro_entries = {}
    
    cedula_ent = ctk.CTkEntry(form_frame, placeholder_text="Cédula", placeholder_text_color=COLOR_PLACEHOLDER, width=ANCHO_INPUT, height=ALTO_INPUT, corner_radius=8, border_width=1, fg_color="white", border_color=COLOR_BORDE, text_color="black", font=ctk.CTkFont(size=14))
    cedula_ent.pack(pady=(0, 12))
    registro_entries['cedula'] = cedula_ent

    nombre_ent = ctk.CTkEntry(form_frame, placeholder_text="Nombre", placeholder_text_color=COLOR_PLACEHOLDER, width=ANCHO_INPUT, height=ALTO_INPUT, corner_radius=8, border_width=1, fg_color="white", border_color=COLOR_BORDE, text_color="black", font=ctk.CTkFont(size=14))
    nombre_ent.pack(pady=(0, 12))
    registro_entries['nombre'] = nombre_ent

    apellido_ent = ctk.CTkEntry(form_frame, placeholder_text="Apellido", placeholder_text_color=COLOR_PLACEHOLDER, width=ANCHO_INPUT, height=ALTO_INPUT, corner_radius=8, border_width=1, fg_color="white", border_color=COLOR_BORDE, text_color="black", font=ctk.CTkFont(size=14))
    apellido_ent.pack(pady=(0, 15))
    registro_entries['apellido'] = apellido_ent

    ctk.CTkLabel(form_frame, text="ROL DE USUARIO", font=ctk.CTkFont(size=12, weight="bold"), text_color="#475569").pack(pady=(5, 2))
    
    rol_vals = rol_names if rol_names else ["-- Sin roles --"]
    rol_combo = ctk.CTkComboBox(form_frame, values=rol_vals, width=ANCHO_INPUT, height=ALTO_INPUT, corner_radius=8, border_width=1, border_color=COLOR_BORDE, fg_color="white", text_color="black", justify="center", button_color="#E2E8F0", button_hover_color="#CBD5E1", dropdown_fg_color="white", dropdown_text_color="black", dropdown_hover_color="#E0F2FE", state="readonly")
    if rol_names:
        default_rol = "usuario" if "usuario" in rol_names else rol_names[0]
        rol_combo.set(default_rol)
    else:
        rol_combo.set("-- Sin roles --")
    rol_combo.pack(pady=(0, 15))
    registro_entries['rol'] = rol_combo

    ctk.CTkLabel(form_frame, text="DEPARTAMENTO", font=ctk.CTkFont(size=12, weight="bold"), text_color="#475569").pack(pady=(5, 2))

    depto_display = ctk.CTkEntry(form_frame, placeholder_text="Seleccione un Departamento...", width=ANCHO_INPUT, height=ALTO_INPUT, corner_radius=8, border_width=1, fg_color="#F8FAFC", border_color=COLOR_BORDE, text_color="black", font=ctk.CTkFont(size=14))
    depto_display.pack(pady=(0, 8))
    
    depto_display.insert(0, departamento_names[0] if departamento_names else "-- Sin departamentos --")
    depto_display.configure(state="readonly")
    
    depto_nombre_var = tk.StringVar(value=departamento_names[0] if departamento_names else "")
    registro_entries['departamento'] = depto_nombre_var

    ctk.CTkButton(form_frame, text="BUSCAR / SELECCIONAR", width=ANCHO_INPUT, height=35, fg_color="#3D89D1", hover_color="#1E3D8F", font=ctk.CTkFont(size=12, weight="bold"), command=lambda: abrir_ventana_seleccion_depto(root, depto_display, depto_nombre_var)).pack(pady=(5, 20))

    btn_limpiar = ctk.CTkButton(form_frame, text="CANCELAR", fg_color="#6B7280", hover_color="#4B5563", font=ctk.CTkFont(size=13, weight="bold"), width=ANCHO_INPUT, height=42, command=lambda: limpiar_formulario())
    btn_limpiar.pack(pady=(0, 10))

    def guardar_usuario():
        cedula_val = (registro_entries.get('cedula').get() or "").strip()
        nombre_val = (registro_entries.get('nombre').get() or "").strip()
        apellido_val = (registro_entries.get('apellido').get() or "").strip()
        rol_nombre = (registro_entries.get('rol').get() or "").strip()
        depto_nombre = (registro_entries.get('departamento').get() or "").strip()

        if not cedula_val or not nombre_val or not apellido_val:
            _set_registro_notificacion("Faltan campos obligatorios.", "orange")
            return
        if not cedula_val.isdigit() or len(cedula_val) < 4:
            _set_registro_notificacion("Cédula inválida o muy corta.", "orange")
            return
        if rol_nombre not in roles_map or depto_nombre not in departamentos_map:
            _set_registro_notificacion("Rol/Departamento no válido.", "red")
            return

        datos_db = {'cedula': int(cedula_val), 'nombre': nombre_val, 'apellido': apellido_val, 'departamento': departamentos_map[depto_nombre], 'rol': roles_map[rol_nombre]}
        datos_visuales = {'cedula': cedula_val, 'nombre': nombre_val.upper(), 'apellido': apellido_val.upper(), 'departamento': depto_nombre, 'rol': rol_nombre}
        _set_registro_notificacion("Procesando...", "#1E3D8F")

        def actualizar_ui_inmediata(es_edicion):
            try:
                if es_edicion and usuario_seleccionado:
                    row = usuario_seleccionado['row_frame']
                    widgets = row.winfo_children()
                    if len(widgets) >= 5:
                        widgets[0].configure(text=datos_visuales['nombre'])
                        widgets[1].configure(text=datos_visuales['apellido'])
                        widgets[2].configure(text=datos_visuales['cedula'])
                        widgets[3].configure(text=datos_visuales['departamento'])
                        widgets[4].configure(text=datos_visuales['rol'])
                    
                    usuario_seleccionado['nombre_completo'] = f"{datos_visuales['nombre']} {datos_visuales['apellido']}"
                    usuario_seleccionado['data'].update(datos_visuales)
                    seleccion_label.configure(text=f"SELECCIONADO: {usuario_seleccionado['nombre_completo']}")
                    _set_registro_notificacion("✓ Usuario actualizado (Vista actualizada)", "#16A34A")
                else:
                    count = len(scroll_frame.winfo_children())
                    bg_color = "#FFFFFF" if count % 2 == 0 else "#F9FAFB"
                    new_row = ctk.CTkFrame(scroll_frame, fg_color=bg_color, corner_radius=0, height=35)
                    new_row.pack(fill="x")
                    new_row.grid_columnconfigure(0, weight=0, minsize=120)
                    new_row.grid_columnconfigure(1, weight=0, minsize=120)
                    new_row.grid_columnconfigure(2, weight=0, minsize=100)
                    new_row.grid_columnconfigure(3, weight=1, minsize=350)
                    new_row.grid_columnconfigure(4, weight=0, minsize=120)
                    
                    font_std = ctk.CTkFont(size=13)
                    color_std = "#374151"
                    lbl_nom = ctk.CTkLabel(new_row, text=datos_visuales['nombre'], font=font_std, text_color=color_std, anchor="w")
                    lbl_ape = ctk.CTkLabel(new_row, text=datos_visuales['apellido'], font=font_std, text_color=color_std, anchor="w")
                    lbl_ced = ctk.CTkLabel(new_row, text=datos_visuales['cedula'], font=font_std, text_color=color_std, anchor="w")
                    lbl_dep = ctk.CTkLabel(new_row, text=datos_visuales['departamento'], font=font_std, text_color=color_std, anchor="w")
                    lbl_rol = ctk.CTkLabel(new_row, text=datos_visuales['rol'], font=font_std, text_color=color_std, anchor="w")
                    
                    lbl_nom.grid(row=0, column=0, padx=8, pady=8, sticky="w")
                    lbl_ape.grid(row=0, column=1, padx=8, pady=8, sticky="w")
                    lbl_ced.grid(row=0, column=2, padx=8, pady=8, sticky="w")
                    lbl_dep.grid(row=0, column=3, padx=8, pady=8, sticky="w")
                    lbl_rol.grid(row=0, column=4, padx=8, pady=8, sticky="w")
                    
                    nombre_completo = f"{datos_visuales['nombre']} {datos_visuales['apellido']}"
                    def bind_click(widget):
                        widget.bind("<Button-1>", lambda e, c=datos_visuales['cedula'], n=nombre_completo, rf=new_row, ud=datos_visuales: seleccionar_usuario(c, n, rf, ud))
                    bind_click(new_row)
                    bind_click(lbl_nom)
                    bind_click(lbl_ape)
                    bind_click(lbl_ced)
                    bind_click(lbl_dep)
                    bind_click(lbl_rol)
                    _set_registro_notificacion("✓ Usuario registrado (Tabla actualizada)", "#16A34A")
                    limpiar_formulario()
            except Exception as e:
                print(f"Error actualizando UI: {e}")
                _set_registro_notificacion("Datos guardados, pero error visual (recargue)", "orange")

        def tarea_guardado():
            try:
                es_edicion = False
                if usuario_seleccionado and str(usuario_seleccionado['cedula']) == str(cedula_val):
                    es_edicion = True
                    resp = supabase.table("Usuario").update(datos_db).eq("cedula", int(cedula_val)).execute()
                else:
                    dup = supabase.table("Usuario").select("cedula").eq("cedula", int(cedula_val)).execute()
                    if dup.data:
                        _set_registro_notificacion("Error: Cédula ya existe.", "red")
                        return
                    resp = supabase.table("Usuario").insert(datos_db).execute()

                if resp.data:
                    app_root.after(0, lambda: actualizar_ui_inmediata(es_edicion))
                else:
                    _set_registro_notificacion("Error al guardar en base de datos.", "red")

            except Exception as e:
                msg = str(e)
                print("Error DB:", msg)
                _set_registro_notificacion("Error de conexión o base de datos.", "red")

        threading.Thread(target=tarea_guardado, daemon=True).start()

    ctk.CTkButton(form_frame, text="GUARDAR USUARIO", fg_color="#16A34A", hover_color="#15803D", font=ctk.CTkFont(size=14, weight="bold"), width=ANCHO_INPUT, height=42, command=guardar_usuario).pack(pady=(0, 6))

    global registro_notificacion 
    registro_notificacion = ctk.CTkLabel(form_frame, text="", font=ctk.CTkFont(size=12, weight="bold"), text_color="#DC2626", wraplength=ANCHO_INPUT)
    registro_notificacion.pack(pady=(5, 15))
    
    def cargar_datos_formulario(usuario_data):
        try:
            cedula_ent.delete(0, 'end')
            cedula_ent.insert(0, usuario_data.get('cedula', ''))
            nombre_ent.delete(0, 'end')
            nombre_ent.insert(0, usuario_data.get('nombre', ''))
            apellido_ent.delete(0, 'end')
            apellido_ent.insert(0, usuario_data.get('apellido', ''))
            
            depto_nombre = usuario_data.get('departamento', '')
            if depto_nombre:
                depto_display.configure(state="normal")
                depto_display.delete(0, 'end')
                depto_display.insert(0, depto_nombre)
                depto_display.configure(state="readonly")
                depto_nombre_var.set(depto_nombre)
            
            rol_nombre = usuario_data.get('rol', '')
            if rol_nombre:
                rol_combo.set(rol_nombre)
            
            registro_notificacion.configure(text="Datos cargados para edición", text_color="#16A34A")
        except Exception as e:
            registro_notificacion.configure(text=f"Error al cargar datos: {e}", text_color="#DC2626")

    def limpiar_formulario():
        global usuario_seleccionado
        usuario_seleccionado = None
        cedula_ent.delete(0, 'end')
        nombre_ent.delete(0, 'end')
        apellido_ent.delete(0, 'end')
        
        if rol_names:
            default_rol = "usuario" if "usuario" in rol_names else rol_names[0]
            rol_combo.set(default_rol)
        
        if departamento_names:
            depto_display.configure(state="normal")
            depto_display.delete(0, 'end')
            depto_display.insert(0, departamento_names[0])
            depto_display.configure(state="readonly")
            depto_nombre_var.set(departamento_names[0])
        
        for widget in scroll_frame.winfo_children():
            if isinstance(widget, ctk.CTkFrame):
                index = scroll_frame.winfo_children().index(widget)
                bg_color = "#FFFFFF" if index % 2 == 0 else "#F9FAFB"
                widget.configure(fg_color=bg_color)
        
        seleccion_label.configure(text="NINGÚN USUARIO SELECCIONADO", text_color="white")
        registro_notificacion.configure(text="Formulario listo para nuevo usuario", text_color="#3D89D1")

    def _eliminar_usuario_seleccionado():
        global usuario_seleccionado
        if not usuario_seleccionado:
            tk.messagebox.showwarning("Advertencia", "Por favor seleccione un usuario primero.")
            return
        row_frame_seleccionado = None
        for widget in scroll_frame.winfo_children():
            if isinstance(widget, ctk.CTkFrame) and widget.cget("fg_color") == "#E0F2FE":
                row_frame_seleccionado = widget
                break
        eliminar_usuario(usuario_seleccionado['cedula'], usuario_seleccionado['nombre_completo'], row_frame_seleccionado)

# PANTALLA PRINCIPAL (Lista de Servicios) 
def mostrar_pantalla_principal(root):
    _clear_widgets(root)
    global app_root
    app_root = root

    filtro_estado = tk.StringVar(value="Todos")
    filtro_fecha = tk.StringVar(value="Todos")
    filtros_especiales = {'tecnico_id': None, 'depto_id': None, 'rango_fecha': None}

    main_frame = ctk.CTkFrame(root, fg_color="#F7F9FB")
    main_frame.pack(expand=True, fill="both")
    main_frame.grid_rowconfigure(1, weight=1)
    main_frame.grid_columnconfigure(0, weight=1)

    header_frame = ctk.CTkFrame(main_frame, fg_color="#0C4A6E", corner_radius=0, height=70)
    header_frame.grid(row=0, column=0, sticky="ew")
    header_frame.grid_columnconfigure(0, weight=1)
    header_frame.grid_columnconfigure(1, weight=0) 
    header_frame.grid_columnconfigure(2, weight=0) 
    header_frame.grid_columnconfigure(3, weight=0) 

    ctk.CTkLabel(header_frame, text="GESTIÓN DE SERVICIOS", font=ctk.CTkFont(size=22, weight="bold"), text_color="white").grid(row=0, column=0, padx=20, pady=15, sticky="w")
                 
    table_card = ctk.CTkFrame(main_frame, fg_color="white", corner_radius=15)
    table_card.grid(row=1, column=0, padx=15, pady=15, sticky="nsew")
    table_card.grid_rowconfigure(1, weight=1) 
    table_card.grid_columnconfigure(0, weight=1)

    title_frame = ctk.CTkFrame(table_card, fg_color="transparent")
    title_frame.grid(row=0, column=0, sticky="ew", padx=18, pady=(15, 5))
    title_frame.grid_columnconfigure(0, weight=1) 

    try:
        logo_img = ctk.CTkImage(PILImage.open("imagen/exportar.png"), size=(200, 60))
        ctk.CTkLabel(title_frame, image=logo_img, text="").grid(row=0, column=0, sticky="w", padx=(10, 0))
    except Exception as e:
        ctk.CTkLabel(title_frame, text="[Logo no encontrado]", text_color="red", font=ctk.CTkFont(size=14, weight="bold")).grid(row=0, column=0, sticky="w")
        
    try:
        reload_icon = ctk.CTkImage(PILImage.open("imagen/recargar.png"), size=(25, 25))
    except Exception:
        reload_icon = None

    try:
        icon_depto = ctk.CTkImage(PILImage.open("imagen/departamento.png"), size=(25, 25))
    except Exception:
        icon_depto = None
    
    try:
        icon_usuario = ctk.CTkImage(PILImage.open("imagen/usuario.png"), size=(25, 25))
    except Exception:
        icon_usuario = None
        
    try:
        icon_sesion = ctk.CTkImage(PILImage.open("imagen/seccion.png"), size=(25, 25))
    except Exception:
        icon_sesion = None

    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        ruta_imagen_boton_exportar = os.path.join(base_dir, "imagen", "btn_exportar.png") 
        export_button_image = ctk.CTkImage(PILImage.open(ruta_imagen_boton_exportar), size=(113, 37))
    except Exception:
        export_button_image = None
        
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        ruta_imagen_graficos_icon = os.path.join(base_dir, "imagen", "grafica.png") 
        icon_graficos = ctk.CTkImage(PILImage.open(ruta_imagen_graficos_icon), size=(25, 25))
    except Exception:
        icon_graficos = None
    
    if icon_depto:
        ctk.CTkButton(header_frame, text="", image=icon_depto, fg_color="#16A34A", hover_color="#15803D", width=40, height=40, corner_radius=8, command=lambda: mostrar_pantalla_departamentos(root)).grid(row=0, column=1, padx=(10, 5), pady=12, sticky="e")
    else:
        ctk.CTkButton(header_frame, text="AGREGAR DEPARTAMENTO", fg_color="#16A34A", hover_color="#15803D", font=ctk.CTkFont(size=13, weight="bold"), corner_radius=8, width=180, height=40, command=lambda: mostrar_pantalla_departamentos(root)).grid(row=0, column=1, padx=(10, 5), pady=12, sticky="e")

    if icon_usuario:
        ctk.CTkButton(header_frame, text="", image=icon_usuario, fg_color="#3D89D1", hover_color="#1E3D8F", width=40, height=40, corner_radius=8, command=lambda: mostrar_pantalla_registro(root)).grid(row=0, column=2, padx=(10, 5), pady=12, sticky="e")
    else:
        ctk.CTkButton(header_frame, text="AGREGAR USUARIO", fg_color="#3D89D1", hover_color="#1E3D8F", font=ctk.CTkFont(size=13, weight="bold"), corner_radius=8, width=140, height=40, command=lambda: mostrar_pantalla_registro(root)).grid(row=0, column=2, padx=(10, 5), pady=12, sticky="e")

    if icon_sesion:
        ctk.CTkButton(header_frame, text="", image=icon_sesion, fg_color="#C82333", hover_color="#A31616", width=40, height=40, corner_radius=8, command=lambda: cerrar_sesion(root)).grid(row=0, column=3, padx=10, pady=12, sticky="e")
    else:
        ctk.CTkButton(header_frame, text="CERRAR SESIÓN", fg_color="#C82333", hover_color="#A31616", command=lambda: cerrar_sesion(root), font=ctk.CTkFont(size=13, weight="bold"), corner_radius=8, width=130, height=40).grid(row=0, column=3, padx=10, pady=12, sticky="e")

    scrollable = ctk.CTkScrollableFrame(table_card, corner_radius=10)
    scrollable.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")

    def obtener_servicios_filtrados():
        query = supabase.table("Servicio").select("*").order("id_servicio", desc=True)
        estado_map = {"Pendiente": 1, "Completado": 2, "Recibido": 3}
        estado_val = filtro_estado.get()
        fecha_val = filtro_fecha.get()
        
        if estado_val == "Todos" and fecha_val == "Todos": 
            query = query.limit(100) 

        if estado_val in estado_map:
            query = query.eq("estado", estado_map[estado_val])
        
        tecnico_id_val = filtros_especiales.get('tecnico_id')
        depto_id_val = filtros_especiales.get('depto_id')

        if tecnico_id_val:
            query = query.eq("tecnico", tecnico_id_val)
        elif depto_id_val:
            departamentos_map = obtener_departamentos()
            nombre_depto = None
            for nombre, idd in departamentos_map.items():
                if idd == depto_id_val:
                    nombre_depto = nombre.strip().lower()
                    break
            if nombre_depto:
                try:
                    query = query.or_(f"departamento.ilike.%{nombre_depto}%,departamento.eq.{depto_id_val}")
                except Exception:
                    query = query.ilike("departamento", f"%{nombre_depto}%")

        hoy_date = datetime.now().date()
        
        if fecha_val == "Hoy":
            inicio_str = hoy_date.isoformat() 
            fin_date = hoy_date + timedelta(days=1)
            fin_str = fin_date.isoformat()
            query = query.gte("fecha", inicio_str).lt("fecha", fin_str)
            
        elif fecha_val == "Ayer":
            ayer_date = hoy_date - timedelta(days=1)
            inicio_str = ayer_date.isoformat()
            fin_date = hoy_date
            fin_str = fin_date.isoformat()
            query = query.gte("fecha", inicio_str).lt("fecha", fin_str)

        elif fecha_val == "Semana anterior":
            inicio_esta_semana = hoy_date - timedelta(days=hoy_date.weekday())
            inicio_semana_anterior = inicio_esta_semana - timedelta(days=7)
            fin_semana_anterior = inicio_esta_semana
            inicio_str = inicio_semana_anterior.isoformat()
            fin_str = fin_semana_anterior.isoformat()
            query = query.gte("fecha", inicio_str).lt("fecha", fin_str)
            
        elif fecha_val == "Personalizado":
            rango = filtros_especiales.get('rango_fecha')
            if rango:
                desde_iso, hasta_iso = rango
                query = query.gte("fecha", desde_iso).lt("fecha", hasta_iso)
            else:
                print("Advertencia: Se seleccionó personalizado pero no hay rango guardado.")

        return obtener_servicios_filtrados_base(query)

    def renderizar_servicios():
        for w in scrollable.winfo_children():
            w.destroy()
        cargando_lbl = ctk.CTkLabel(scrollable, text="Consultando servicios...", font=ctk.CTkFont(size=14, weight="bold"), text_color="#0C4A6E")
        cargando_lbl.pack(pady=20)
        
        estado_paginacion = {"datos": [], "usuarios_map": {}, "indice_actual": 0, "lote_tamano": 10, "btn_cargar_mas": None}

        def tarea_obtener_datos():
            try:
                raw_servicios = obtener_servicios_filtrados() 
                estado_paginacion["datos"] = raw_servicios
                estado_paginacion["usuarios_map"] = map_usuarios_por_cedula()
            except Exception as e:
                print(f"Error obteniendo datos: {e}")
                estado_paginacion["datos"] = []
                estado_paginacion["usuarios_map"] = {}
            scrollable.after(0, iniciar_vista)

        def iniciar_vista():
            try:
                cargando_lbl.destroy()
            except:
                pass
            if not estado_paginacion["datos"]:
                ctk.CTkLabel(scrollable, text="No hay servicios registrados con estos filtros.", font=ctk.CTkFont(size=14), text_color="gray").pack(pady=20)
                return
            scrollable._parent_canvas.yview_moveto(0.0)
            scrollable.grid_columnconfigure(0, weight=1)
            cargar_siguiente_lote()

        def cargar_siguiente_lote():
            if estado_paginacion["btn_cargar_mas"]:
                estado_paginacion["btn_cargar_mas"].destroy()
                estado_paginacion["btn_cargar_mas"] = None

            inicio = estado_paginacion["indice_actual"]
            fin = inicio + estado_paginacion["lote_tamano"]
            lista_completa = estado_paginacion["datos"]
            lote = lista_completa[inicio:fin]

            COLOR_HEADER_BG = "#0A2B4C"
            COLOR_BODY_BG = "#F5F5ED"
            COLOR_HEADER_TEXT = "#FFFFFF"
            COLOR_DETAIL_TEXT = "#4A4A4A"
            COLOR_SEPARATOR = "#DCDCDC" 
            CARD_CORNER_RADIUS = 8
            
            FONT_HEADER = ctk.CTkFont(size=18, weight="bold")
            FONT_DETAIL = ctk.CTkFont(size=15)
            FONT_PILL = ctk.CTkFont(size=11, weight="bold")

            colores_estado = {
                "Completado": ("#D1FAE5", "#047857", "#047857"),
                "Pendiente":  ("#FEF3C7", "#92400E", "#92400E"),
                "Recibido":   ("#DBEAFE", "#1E3A8A", "#1E3A8A"),
                "Desconocido": ("#F3F4F6", "#374151", "#374151")
            }
            
            col_min_width = 300
            wrap_width = col_min_width - 20

            for i, s in enumerate(lote):
                grid_index = inicio + i
                
                estado_text = traducir_estado(s.get("estado"))
                color_bg, color_border, color_text = colores_estado.get(estado_text, colores_estado["Desconocido"])
                
                titulo_val = (s.get('descripcion') or "Sin descripción").capitalize()
                usuario_val = estado_paginacion["usuarios_map"].get(str(s.get('usuario')), 'Desconocido')
                depto_val = s.get('Departamento', 'Desconocido')
                tecnico_val = estado_paginacion["usuarios_map"].get(str(s.get('tecnico')), 'Sin asignar')
                
                reporte_valor = s.get("reporte")
                if not reporte_valor or str(reporte_valor).strip().lower() in ["none", "null", ""]:
                    reporte_valor = "Sin reporte"

                card_main = ctk.CTkFrame(scrollable, fg_color=COLOR_BODY_BG, corner_radius=CARD_CORNER_RADIUS, border_color="#DCDCDC", border_width=1)
                card_main.grid(row=grid_index, column=0, sticky="ew", padx=15, pady=8)
                card_main.grid_columnconfigure(0, weight=1)

                header_frame = ctk.CTkFrame(card_main, fg_color=COLOR_HEADER_BG, corner_radius=0)
                header_frame.grid(row=0, column=0, sticky="ew")
                ctk.CTkLabel(header_frame, text=f" SERVICIO #{s.get('id_servicio')} | {titulo_val}", font=FONT_HEADER, text_color=COLOR_HEADER_TEXT, anchor="w").pack(fill="x", padx=15, pady=10)

                body_container = ctk.CTkFrame(card_main, fg_color="transparent")
                body_container.grid(row=1, column=0, sticky="nsew", padx=15, pady=5)
                body_container.grid_columnconfigure(0, weight=1)
                body_container.grid_rowconfigure(0, weight=1)

                columns_frame = ctk.CTkFrame(body_container, fg_color="transparent")
                columns_frame.grid(row=0, column=0, sticky="nsew")
                columns_frame.grid_columnconfigure(0, weight=1, minsize=col_min_width)
                columns_frame.grid_columnconfigure(2, weight=1, minsize=col_min_width)
                columns_frame.grid_columnconfigure(4, weight=1, minsize=col_min_width)
                columns_frame.grid_columnconfigure((1, 3), weight=0)

                c1 = ctk.CTkFrame(columns_frame, fg_color="transparent")
                c1.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
                ctk.CTkLabel(c1, text=f"Usuario: {usuario_val}", font=FONT_DETAIL, text_color=COLOR_DETAIL_TEXT, anchor="w", justify="left", wraplength=wrap_width).pack(fill="x", anchor="w")
                ctk.CTkLabel(c1, text=f"Departamento: {depto_val}", font=FONT_DETAIL, text_color=COLOR_DETAIL_TEXT, anchor="w", justify="left", wraplength=wrap_width).pack(fill="x", anchor="w")

                ctk.CTkFrame(columns_frame, width=2, fg_color=COLOR_SEPARATOR).grid(row=0, column=1, sticky="ns", pady=5)

                c2 = ctk.CTkFrame(columns_frame, fg_color="transparent")
                c2.grid(row=0, column=2, sticky="nsew", padx=5)
                ctk.CTkLabel(c2, text=f"Técnico: {tecnico_val}", font=FONT_DETAIL, text_color=COLOR_DETAIL_TEXT, anchor="w", justify="left", wraplength=wrap_width).pack(fill="x", anchor="w")
                ctk.CTkLabel(c2, text=f"Reporte: {reporte_valor}", font=FONT_DETAIL, text_color=COLOR_DETAIL_TEXT, anchor="w", justify="left", wraplength=wrap_width).pack(fill="x", anchor="w")

                ctk.CTkFrame(columns_frame, width=2, fg_color=COLOR_SEPARATOR).grid(row=0, column=3, sticky="ns", pady=5)

                c3 = ctk.CTkFrame(columns_frame, fg_color="transparent")
                c3.grid(row=0, column=4, sticky="nsew", padx=(5, 0))
                ctk.CTkLabel(c3, text=f"Fecha creación: {formatear_fecha(s.get('fecha'))}", font=FONT_DETAIL, text_color=COLOR_DETAIL_TEXT, anchor="w").pack(fill="x", anchor="w")
                ctk.CTkLabel(c3, text=f"Fecha de culminación: {formatear_fecha(s.get('fecha_culminado'))}", font=FONT_DETAIL, text_color=COLOR_DETAIL_TEXT, anchor="w").pack(fill="x", anchor="w")

                pill = ctk.CTkFrame(body_container, fg_color=color_bg, border_color=color_border, border_width=1, corner_radius=14)
                pill.grid(row=0, column=0, padx=(0, 5), pady=(0, 5), sticky="se")
                ctk.CTkLabel(pill, text=estado_text.upper(), text_color=color_text, font=FONT_PILL).pack(padx=12, pady=5)

            estado_paginacion["indice_actual"] = fin
            registros_restantes = len(lista_completa) - estado_paginacion["indice_actual"]
            
            if registros_restantes > 0:
                btn = ctk.CTkButton(scrollable, text=f"▼ MOSTRAR MÁS ({registros_restantes} restantes) ▼", fg_color="#E0E7FF", text_color="#1E3A8A", hover_color="#C7D2FE", font=ctk.CTkFont(size=12, weight="bold"), height=35, command=lambda: cargar_siguiente_lote())
                btn.grid(row=fin + 1, column=0, pady=15, sticky="ew", padx=40)
                estado_paginacion["btn_cargar_mas"] = btn
            else:
                lbl_fin = ctk.CTkLabel(scrollable, text="--- No hay más registros ---", text_color="gray", font=ctk.CTkFont(size=11))
                lbl_fin.grid(row=fin + 1, column=0, pady=20)

        threading.Thread(target=tarea_obtener_datos, daemon=True).start()
    
    def exportar_a_excel():
        def tarea_exportar():
            try:
                servicios = obtener_servicios_filtrados()
                usuarios_map = map_usuarios_por_cedula()
                if not servicios:
                    root.after(0, lambda: messagebox.showwarning("Sin datos", "No hay servicios filtrados para exportar."))
                    return
                datos_para_excel = []
                columnas_excel = ['ID Servicio', 'Estado', 'Descripción', 'Usuario', 'Técnico', 'Departamento', 'Fecha Creación', 'Reporte', 'Fecha Culminado']
                
                for s in servicios:
                    estado_text = traducir_estado(s.get("estado"))
                    reporte_valor = s.get("reporte")
                    if not reporte_valor or str(reporte_valor).strip().lower() in ["none", "null", ""]:
                        reporte_valor = "No registrado"
                    usuario_nombre = usuarios_map.get(str(s.get('usuario')), 'Desconocido')
                    tecnico_nombre = usuarios_map.get(str(s.get('tecnico')), 'Sin asignar')
                    depto_nombre = s.get('Departamento', 'Desconocido')
                    fila = [s.get('id_servicio'), estado_text, s.get('descripcion'), usuario_nombre, tecnico_nombre, depto_nombre, formatear_fecha(s.get('fecha')), reporte_valor, formatear_fecha(s.get('fecha_culminado'))]
                    datos_para_excel.append(fila)
                df = pd.DataFrame(datos_para_excel, columns=columnas_excel)
                root.after(0, lambda: _guardar_excel_en_hilo_principal(df))
            except Exception as e:
                print(f"Error en hilo de exportación: {e}")
                root.after(0, lambda: messagebox.showerror("Error", f"Ocurrió un error al preparar los datos:\n{e}"))
     
        def _guardar_excel_en_hilo_principal(df):
            try:
                ruta_archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")], title="Guardar reporte de servicios")
                if not ruta_archivo:
                    return
                wb = Workbook()
                ws = wb.active
                ws.title = "Servicios"

                try:
                    base_dir = os.path.dirname(os.path.abspath(__file__))
                    ruta_imagen = os.path.join(base_dir, "imagen", "exportar.png")
                    if os.path.exists(ruta_imagen):
                        img = OpenpyxlImage(ruta_imagen)
                        img.width = 290
                        img.height = 73
                        ws.add_image(img, "A1")
                        ws.column_dimensions['A'].width = 10
                        ws.column_dimensions['B'].width = 20
                        ws.column_dimensions['C'].width = 2
                        fila_titulo = 1
                    else:
                        fila_titulo = 1
                except Exception:
                    fila_titulo = 1
                    
                ultima_columna_letra = get_column_letter(df.shape[1])
                ws.merge_cells(f'C{fila_titulo}:{ultima_columna_letra}{fila_titulo}')
                ws[f'C{fila_titulo}'] = "GESTIÓN DE SERVICIOS DGTSP\nARAGUA"
                ws[f'C{fila_titulo}'].font = Font(name='Arial', size=14, bold=True, color="000000")
                ws[f'C{fila_titulo}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.row_dimensions[1].height = 45

                for col in ['A', 'B', 'C']:
                    cell = ws[f"{col}1"]
                    cell.border = Border(left=None, right=None, top=None, bottom=None)
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                ws.sheet_view.showGridLines = False 

                header_row = fila_titulo + 2
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=header_row, column=col_idx, value=col_name)
                    cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(left=Side(style=BORDER_THIN), right=Side(style=BORDER_THIN), top=Side(style=BORDER_THIN), bottom=Side(style=BORDER_THIN))

                for r_idx, row in enumerate(df.values, header_row + 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.font = Font(name='Arial', size=9, color="000000")
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                        cell.border = Border(left=Side(style=BORDER_THIN), right=Side(style=BORDER_THIN), top=Side(style=BORDER_THIN), bottom=Side(style=BORDER_THIN))
                        
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column[3:]: 
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    if column_letter == 'A': ws.column_dimensions[column_letter].width = max(adjusted_width, 12)
                    elif column_letter == 'B': ws.column_dimensions[column_letter].width = min(max(adjusted_width, 13), 18)
                    elif column_letter == 'C': ws.column_dimensions[column_letter].width = max(adjusted_width, 40)
                    elif column_letter == 'D': ws.column_dimensions[column_letter].width = max(adjusted_width, 20)
                    elif column_letter == 'E': ws.column_dimensions[column_letter].width = max(adjusted_width, 20)
                    elif column_letter == 'F': ws.column_dimensions[column_letter].width = max(adjusted_width, 25)
                    elif column_letter == 'G': ws.column_dimensions[column_letter].width = max(adjusted_width, 18)
                    elif column_letter == 'H': ws.column_dimensions[column_letter].width = max(adjusted_width, 30)
                    elif column_letter == 'I': ws.column_dimensions[column_letter].width = max(adjusted_width, 18)
                    else: ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(ruta_archivo)
                os.startfile(ruta_archivo)
            except Exception as e:
                messagebox.showerror("Error al guardar", f"No se pudo guardar el archivo:\n{e}")
        threading.Thread(target=tarea_exportar, daemon=True).start()

    def abrir_ventana_filtrar_departamento():
        ventana = ctk.CTkToplevel(root)
        ventana.title("Filtrar por Departamento")
        ventana.configure(fg_color="#F7F9FB")
        ventana.geometry("580x600") 
        ventana.grab_set()
        ventana.focus_force()
        ventana.resizable(False, False)
        
        seleccion_temp = {"id": None, "nombre": None, "widget": None}
        contenido = ctk.CTkFrame(ventana, fg_color="#FFFFFF")
        contenido.pack(padx=20, pady=20, fill="both", expand=True)
        contenido.grid_columnconfigure(0, weight=1)
        contenido.grid_rowconfigure(2, weight=1) 

        ctk.CTkLabel(contenido, text="Buscar Departamento", font=ctk.CTkFont(size=18, weight="bold"), text_color="#0C4A6E").grid(row=0, column=0, pady=(15, 10), sticky="w", padx=20)
        search_entry = ctk.CTkEntry(contenido, placeholder_text="Escriba la inicial...", height=40, border_color="#A1A1A1")
        search_entry.grid(row=1, column=0, pady=(0, 15), padx=20, sticky="ew")
        
        scroll_frame = ctk.CTkScrollableFrame(contenido, fg_color="#F9FAFB")
        scroll_frame.grid(row=2, column=0, sticky="nsew", padx=20, pady=(0, 10)) 
        scroll_frame.grid_columnconfigure(0, weight=1)

        deptos_map = obtener_departamentos() 
        all_nombres = sorted(list(deptos_map.keys()))

        def seleccionar_item(nombre, idd, btn_widget):
            if seleccion_temp["widget"] and seleccion_temp["widget"].winfo_exists():
                seleccion_temp["widget"].configure(fg_color="transparent", text_color="black")
            seleccion_temp["id"] = idd
            seleccion_temp["nombre"] = nombre
            seleccion_temp["widget"] = btn_widget
            btn_widget.configure(fg_color="#0C4A6E", text_color="white")

        def render_lista(filtro=""):
            for w in scroll_frame.winfo_children():
                w.destroy()
            texto_busqueda = filtro.lower().strip()
            encontrados = False
            for nombre in all_nombres:
                if not texto_busqueda or nombre.lower().startswith(texto_busqueda):
                    encontrados = True
                    idd = deptos_map[nombre]
                    es_seleccionado = (idd == seleccion_temp["id"])
                    bg_color = "#0C4A6E" if es_seleccionado else "transparent"
                    fg_txt = "white" if es_seleccionado else "black"

                    btn = ctk.CTkButton(scroll_frame, text=nombre, fg_color=bg_color, text_color=fg_txt, hover_color="#3D89D1", anchor="w", height=40, font=ctk.CTkFont(size=13), command=None)
                    btn.configure(command=lambda n=nombre, i=idd, b=btn: seleccionar_item(n, i, b))
                    btn.pack(fill="x", pady=1)
                    if es_seleccionado:
                        seleccion_temp["widget"] = btn
            if not encontrados:
                ctk.CTkLabel(scroll_frame, text="No hay departamentos con esa inicial", text_color="gray").pack(pady=20)

        def ejecutar_filtro():
            if seleccion_temp["id"] is None:
                messagebox.showwarning("Atención", "Por favor, seleccione un departamento.")
                return
            filtros_especiales['depto_id'] = seleccion_temp["id"]
            filtros_especiales['tecnico_id'] = None 
            nombre_corto = seleccion_temp["nombre"]
            if len(nombre_corto) > 20: 
                nombre_corto = nombre_corto[:20] + "..."
            filtro_estado.set(f"Depto: {nombre_corto}")
            ventana.destroy()
            renderizar_servicios()

        btn_aplicar = ctk.CTkButton(contenido, text="APLICAR FILTRO", fg_color="#0C4A6E", hover_color="#155E75", height=45, font=ctk.CTkFont(size=14, weight="bold"), command=ejecutar_filtro)
        btn_aplicar.grid(row=3, column=0, pady=20, padx=20, sticky="ew")
        search_entry.bind("<KeyRelease>", lambda event: render_lista(search_entry.get()))
        render_lista()

    def manejar_filtro_fecha(opcion):
        if opcion == "Personalizado":
            try:
                from tkcalendar import Calendar
            except ImportError:
                messagebox.showerror("Error", "La librería 'tkcalendar' no está instalada.\nEjecute: pip install tkcalendar")
                filtro_fecha.set("Todos") 
                renderizar_servicios()
                return

            ventana = ctk.CTkToplevel(root)
            ventana.title("Seleccionar rango de fechas")
            ventana.configure(fg_color="#F7F9FB")
            ventana.geometry("650x420") 
            ventana.resizable(False, False)
            ventana.grab_set()
            ventana.focus_force()

            contenido = ctk.CTkFrame(ventana, fg_color="#F7F9FB")
            contenido.pack(padx=20, pady=20, fill="both", expand=True)
            contenido.grid_columnconfigure(0, weight=1)
            contenido.grid_columnconfigure(1, weight=1)
            contenido.grid_rowconfigure(2, weight=1) 

            ctk.CTkLabel(contenido, text="Seleccione el rango de fechas", font=ctk.CTkFont(size=18, weight="bold"), text_color="#0C4A6E").grid(row=0, column=0, columnspan=2, pady=(10, 15))
            ctk.CTkLabel(contenido, text="Desde:", text_color="#2E3A59", font=ctk.CTkFont(size=13, weight="bold")).grid(row=1, column=0, pady=(5, 5))
            ctk.CTkLabel(contenido, text="Hasta:", text_color="#2E3A59", font=ctk.CTkFont(size=13, weight="bold")).grid(row=1, column=1, pady=(5, 5))

            cal_desde = Calendar(contenido, date_pattern="dd-mm-yyyy", selectmode="day")
            cal_hasta = Calendar(contenido, date_pattern="dd-mm-yyyy", selectmode="day")
            cal_desde.grid(row=2, column=0, padx=15, pady=(0, 10))
            cal_hasta.grid(row=2, column=1, padx=15, pady=(0, 10))

            def aplicar():
                desde_str, hasta_str = cal_desde.get_date(), cal_hasta.get_date()
                try:
                    desde_obj = datetime.strptime(desde_str, "%d-%m-%Y").date()
                    hasta_obj = datetime.strptime(hasta_str, "%d-%m-%Y").date()
                except ValueError:
                    return
                if desde_obj > hasta_obj:
                    messagebox.showwarning("Fechas inválidas", "La fecha 'Desde' no puede ser mayor que 'Hasta'.")
                    return
                hasta_inclusive = hasta_obj + timedelta(days=1)
                filtros_especiales['rango_fecha'] = (desde_obj.isoformat(), hasta_inclusive.isoformat())
                ventana.destroy()
                renderizar_servicios()

            btn_aplicar = ctk.CTkButton(contenido, text="APLICAR FILTRO", fg_color="#0C4A6E", hover_color="#155E75", corner_radius=10, width=200, height=40, command=aplicar)
            btn_aplicar.grid(row=3, column=0, columnspan=2, pady=(20, 10))
        else:
            renderizar_servicios()

    def manejar_filtro_principal(opcion):
        if opcion == "Por Departamento...":
            abrir_ventana_filtrar_departamento()      
            
        else:
            filtros_especiales['tecnico_id'] = None
            filtros_especiales['depto_id'] = None
            renderizar_servicios()
            
    filtro_estado_menu = ctk.CTkOptionMenu(title_frame, values=["Todos", "Pendiente", "Recibido", "Completado", "Por Departamento..."], variable=filtro_estado, command=manejar_filtro_principal, fg_color="#0C4A6E", button_color="#155E75", text_color="white", width=200, height=35, dropdown_fg_color="#E5E7EB", dropdown_text_color="black")
    filtro_estado_menu.grid(row=0, column=1, padx=5, sticky="e")

    filtro_fecha_menu = ctk.CTkOptionMenu(title_frame, values=["Todos", "Hoy", "Ayer", "Semana anterior", "Personalizado"], variable=filtro_fecha, command=manejar_filtro_fecha, fg_color="#0C4A6E", button_color="#155E75", text_color="white", width=180, height=35, dropdown_fg_color="#E5E7EB", dropdown_text_color="black")
    filtro_fecha_menu.grid(row=0, column=2, padx=5, sticky="e")

    if export_button_image: 
        ctk.CTkButton(title_frame, text="", image=export_button_image, width=100, height=35, fg_color="transparent", hover_color="#C3DBB9", command=exportar_a_excel).grid(row=0, column=3, padx=4, sticky="e")
    else: 
         ctk.CTkButton(title_frame, text="Exportar", width=100, height=35, fg_color="#107C41", hover_color="#0B532B", corner_radius=8, command=exportar_a_excel).grid(row=0, column=3, padx=5, sticky="e")
    
    if icon_graficos: 
        ctk.CTkButton(title_frame, text="", image=icon_graficos, width=45, height=35, fg_color="#D97706", hover_color="#B45309", corner_radius=8, command=lambda: mostrar_pantalla_graficos(root, mostrar_pantalla_principal)).grid(row=0, column=4, padx=5, sticky="e")
    else:
        ctk.CTkButton(title_frame, text="Gráficos", width=100, height=35, fg_color="#D97706", hover_color="#B45309", corner_radius=8, command=lambda: mostrar_pantalla_graficos(root, mostrar_pantalla_principal)).grid(row=0, column=4, padx=5, sticky="e")

    ctk.CTkButton(title_frame, text="", image=reload_icon, width=45, height=35, fg_color="#E5E7EB", hover_color="#CBD5E1", corner_radius=8, command=renderizar_servicios).grid(row=0, column=5, padx=5, sticky="e") 
    
    renderizar_servicios()